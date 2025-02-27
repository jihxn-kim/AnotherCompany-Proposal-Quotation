from openpyxl import load_workbook
import time
import os

class Sheet:
    def __init__(self, class_num, directory_path, class_date, save_path,
                first_program_1, first_program_2, first_grade, first_class,
                second_program_1, second_program_2, second_grade, second_class,
                last_program_1, last_program_2, last_grade, last_class,
                school_name, price):
        self.first_grade = first_grade
        self.first_class = first_class
        self.class_num = class_num
        self.directory_path = directory_path
        self.save_path = save_path
        self.class_date = class_date
        self.first_program_1 = first_program_1
        self.first_program_2 = first_program_2
        self.second_program_1 = second_program_1
        self.second_program_2 = second_program_2
        self.second_grade =second_grade
        self.second_class = second_class
        self.last_program_1 = last_program_1
        self.last_program_2 = last_program_2
        self.last_grade = last_grade
        self.last_class = last_class
        self.school_name = school_name
        self.price = price

    def makeSheet(self):
        sampleXlsx = self.directory_path + "/견적서/threeProgram.xlsx"
        
        wb = load_workbook(sampleXlsx)
        ws = wb.active

        ## 학교명
        ws["H3"] = self.school_name 

        ## 사업명
        ws["H4"] = "시뮬레이션 진로캠프"

        ## 견적일자
        curr_time = time.strftime("%Y-%m-%d").strip()
        ws["H7"] = curr_time

        ## 강사비
        # 내용
        ws["B10"] = self.first_grade + "강사비"
        ws["B11"] = self.second_grade + "강사비"
        ws["B12"] = self.last_grade + "강사비"

        # 단가 (50,000 고정)

        # 수량
        ws["D10"] = self.first_class
        ws["D11"] = self.second_class
        ws["D12"] = self.last_class

        # 회
        ws["F10"] = self.class_num[:1]
        ws["F11"] = self.class_num[:1]
        ws["F12"] = self.class_num[:1]

        # 합계
        ws["G10"] = 50000 * int(self.class_num[:1]) * int(self.first_class)
        ws["G11"] = 50000 * int(self.class_num[:1]) * int(self.second_class)
        ws["G12"] = 50000 * int(self.class_num[:1]) * int(self.last_class)

        # 비고
        ws["H10"] = self.class_num
        ws["H11"] = self.class_num
        ws["H12"] = self.class_num

        ## 재료비
        # 내용
        ws["B13"] = self.first_grade + "재료비"
        ws["B14"] = self.second_grade + "재료비"
        ws["B15"] = self.last_grade + "재료비"

        # 단가
        material_cost_1 = int(self.price) - 50000 * int(self.class_num[:1])
        ws["C13"] = material_cost_1
        ws["C14"] = material_cost_1
        ws["C15"] = material_cost_1

        # 수량
        ws["D13"] = self.first_class
        ws["D14"] = self.second_class
        ws["D15"] = self.last_class

        # 합계
        ws["G13"] = material_cost_1 * int(self.first_class)
        ws["G14"] = material_cost_1 * int(self.second_class)
        ws["G15"] = material_cost_1 * int(self.last_class)

        # 비고
        if self.first_program_1 == "수상한스튜디오":
            material_1 = "수상한스튜디오 툴킷(비전욕망카드 1set, 숫자카드), 교재(A4, 4P, 컬러)"
        elif self.first_program_1 == "어나더랜드":
            material_1 = "어나더랜드 툴킷(손목밴드, 어나더게임 카드 외), 교재(A4, 8P, 컬러)"
        elif self.first_program_1 == "취업조작단":
            material_1 = "취업조작단 툴킷(강점스캐닝 교구(스티커 1set)), 교재(A4, 양면)"
        elif self.first_program_1 == "비밀상담소":
            material_1 = "고교학점제 툴킷(고교학점제 카드), 교재 (A4, 8P, 컬러)"
        elif self.first_program_1 == "코드5":
            material_1 = "CODE 5 툴킷(직무카드 330장, 플레이매트, 점수카드 등), 교재 (A3, 컬러, 2매)"
        elif self.first_program_1 == "AI오피스":
            pass

        if self.second_program_1 == "수상한스튜디오":
            material_2 = "수상한스튜디오 툴킷(비전욕망카드 1set, 숫자카드), 교재(A4, 4P, 컬러)"
        elif self.second_program_1 == "어나더랜드":
            material_2 = "어나더랜드 툴킷(손목밴드, 어나더게임 카드 외), 교재(A4, 8P, 컬러)"
        elif self.second_program_1 == "취업조작단":
            material_2 = "취업조작단 툴킷(강점스캐닝 교구(스티커 1set)), 교재(A4, 양면)"
        elif self.second_program_1 == "비밀상담소":
            material_2 = "고교학점제 툴킷(고교학점제 카드), 교재 (A4, 8P, 컬러)"
        elif self.second_program_1 == "코드5":
            material_2 = "CODE 5 툴킷(직무카드 330장, 플레이매트, 점수카드 등), 교재 (A3, 컬러, 2매)"
        elif self.second_program_1 == "AI오피스":
            pass

        if self.last_program_1 == "수상한스튜디오":
            material_3 = "수상한스튜디오 툴킷(비전욕망카드 1set, 숫자카드), 교재(A4, 4P, 컬러)"
        elif self.last_program_1 == "어나더랜드":
            material_3 = "어나더랜드 툴킷(손목밴드, 어나더게임 카드 외), 교재(A4, 8P, 컬러)"
        elif self.last_program_1 == "취업조작단":
            material_3 = "취업조작단 툴킷(강점스캐닝 교구(스티커 1set)), 교재(A4, 양면)"
        elif self.last_program_1 == "비밀상담소":
            material_3 = "고교학점제 툴킷(고교학점제 카드), 교재 (A4, 8P, 컬러)"
        elif self.last_program_1 == "코드5":
            material_3 = "CODE 5 툴킷(직무카드 330장, 플레이매트, 점수카드 등), 교재 (A3, 컬러, 2매)"
        elif self.last_program_1 == "AI오피스":
            pass

        ws["H13"] = material_1
        ws["H14"] = material_2
        ws["H15"] = material_3

        ## 전체 합계
        ws["G16"] = self.price * (self.first_class + self.second_class + self.last_class)

        ## 파일명
        file_name = self.save_path + "/" + time.strftime("%y%m%d").strip() + "_" + self.school_name + "_시뮬레이션진로캠프_견적서_어나더컴퍼니.xlsx"
        wb.save(file_name)

        absolute_path = os.path.join(os.getcwd(), file_name)
        return absolute_path


        