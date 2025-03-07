from openpyxl import load_workbook
from openpyxl.styles import Alignment
import time
import os

class Sheet:
    def __init__(self, first_grade, first_class, class_num, directory_path, save_path,
                 first_program_1, first_program_2,
                 school_name, price):
        self.first_grade = first_grade
        self.first_class = first_class
        self.class_num = class_num
        self.directory_path = directory_path
        self.save_path = save_path
        self.first_program_1 = first_program_1
        self.first_program_2 = first_program_2
        self.school_name = school_name
        self.price = price

    def makeSheet(self):
        sampleXlsx = self.directory_path + "/견적서/oneProgram.xlsx"
        
        wb = load_workbook(sampleXlsx)
        ws = wb.active

        ## 학교명
        ws["B8"] = self.school_name 

        ## 견적일자
        curr_time = time.strftime("%Y-%m-%d").strip()
        ws["B10"] = curr_time

        # 내용
        ws["A14"] = self.first_program_1 + "\n" + "프로그램 체험비"
        ws["A14"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

        # 차시
        ws["B14"] = self.class_num

        # 단가
        ws["C14"] = int(self.price)
        ws["C14"].number_format = '"₩"#,##0'

        # 수량
        ws["E14"] = self.first_class

        # 합계
        ws["H14"] = int(self.price) * int(self.first_class)
        ws["H14"].number_format = '"₩"#,##0'

        # 견적가 총합
        ws["B15"] = int(self.price) * int(self.first_class)
        ws["B15"].number_format = '"₩"#,##0'

        ## 파일명
        curr_time = time.strftime("%y%m%d").strip()
        curr_year = time.strftime("%Y") + "년"
        curr_month = time.strftime("%m") + "월"

        save_dir = os.path.join(self.save_path, curr_year, curr_month, self.school_name)
        os.makedirs(save_dir, exist_ok=True)

        file_name = f"{curr_time}_{self.school_name}_{self.first_grade}_{self.first_program_1}_견적서_어나더컴퍼니.xlsx"
        file_path = os.path.join(save_dir, file_name)

        wb.save(file_path)

        absolute_path = os.path.join(os.getcwd(), file_path)
        print("파일경로: " + absolute_path)
        return absolute_path