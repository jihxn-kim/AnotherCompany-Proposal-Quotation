from openpyxl import load_workbook
from openpyxl.styles import Alignment
import time
import os

class Sheet:
    def __init__(self, directory_path, save_path,
                first_program_1, first_program_2, first_grade, first_class, class_num_1, price_1,
                second_program_1, second_program_2, second_grade, second_class, class_num_2, price_2,
                school_name):
        self.directory_path = directory_path
        self.save_path = save_path

        self.first_grade = first_grade
        self.first_class = first_class
        self.first_program_1 = first_program_1
        self.first_program_2 = first_program_2
        self.class_num_1 = class_num_1
        self.price_1 = price_1

        self.second_program_1 = second_program_1
        self.second_program_2 = second_program_2
        self.second_grade =second_grade
        self.second_class = second_class
        self.class_num_2 = class_num_2
        self.price_2 = price_2

        self.school_name = school_name

    def makeSheet(self):
        sampleXlsx = self.directory_path + "/견적서/twoProgram.xlsx"
        
        wb = load_workbook(sampleXlsx)
        ws = wb.active

        ## 학교명
        ws["B8"] = self.school_name 

        ## 견적일자
        curr_time = time.strftime("%Y-%m-%d").strip()
        ws["B10"] = curr_time

        # 내용
        ws["A14"] = self.first_grade + "\n" + self.first_program_1 + "\n" + "프로그램 체험비"
        ws["A14"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

        ws["A15"] = self.second_grade + "\n" + self.second_program_1 + "\n" + "프로그램 체험비"
        ws["A15"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

        # 차시
        ws["B14"] = self.class_num_1
        ws["B15"] = self.class_num_2

        # 단가
        ws["C14"] = int(self.price_1)
        ws["C14"].number_format = '"₩"#,##0'

        ws["C15"] = int(self.price_2)
        ws["C15"].number_format = '"₩"#,##0'

        # 수량
        ws["E14"] = self.first_class
        ws["E15"] = self.second_class

        # 합계
        ws["H14"] = int(self.price_1) * int(self.first_class)
        ws["H14"].number_format = '"₩"#,##0'

        ws["H15"] = int(self.price_2) * int(self.second_class)
        ws["H15"].number_format = '"₩"#,##0'

        # 견적가 총합
        ws["B16"] = int(self.price_1) * int(self.first_class) + int(self.price_2) * int(self.second_class)
        ws["B16"].number_format = '"₩"#,##0'

        ## 파일명
        curr_time = time.strftime("%y%m%d").strip()
        curr_year = time.strftime("%Y") + "년"
        curr_month = time.strftime("%m") + "월"

        save_dir = os.path.join(self.save_path, curr_year, curr_month, self.school_name)
        os.makedirs(save_dir, exist_ok=True)

        file_name = f"{curr_time}_{self.school_name}_시뮬레이션진로캠프_견적서_어나더컴퍼니.xlsx"
        file_path = os.path.join(save_dir, file_name)

        wb.save(file_path)

        absolute_path = os.path.join(os.getcwd(), file_path)
        return absolute_path